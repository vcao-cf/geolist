import argparse, csv, io, json, re, zipfile
from pathlib import Path
from openpyxl import load_workbook

STATE_NAMES=dict(AL="Alabama",AK="Alaska",AZ="Arizona",AR="Arkansas",CA="California",CO="Colorado",CT="Connecticut",DE="Delaware",FL="Florida",GA="Georgia",HI="Hawaii",ID="Idaho",IL="Illinois",IN="Indiana",IA="Iowa",KS="Kansas",KY="Kentucky",LA="Louisiana",ME="Maine",MD="Maryland",MA="Massachusetts",MI="Michigan",MN="Minnesota",MS="Mississippi",MO="Missouri",MT="Montana",NE="Nebraska",NV="Nevada",NH="New Hampshire",NJ="New Jersey",NM="New Mexico",NY="New York",NC="North Carolina",ND="North Dakota",OH="Ohio",OK="Oklahoma",OR="Oregon",PA="Pennsylvania",RI="Rhode Island",SC="South Carolina",SD="South Dakota",TN="Tennessee",TX="Texas",UT="Utah",VT="Vermont",VA="Virginia",WA="Washington",WV="West Virginia",WI="Wisconsin",WY="Wyoming",DC="District of Columbia",PR="Puerto Rico")

def norm(v): return re.sub(r"[^A-Z0-9]","",str(v or "").upper())
def rows(path):
    if path.suffix.lower() in (".xlsx",".xlsm"):
        sheet=load_workbook(path,read_only=True,data_only=True).active
        it=sheet.iter_rows(values_only=True); headers=[norm(v) for v in next(it)]
        for values in it: yield dict(zip(headers,values))
    else:
        with path.open("r",encoding="utf-8-sig",newline="") as f:
            reader=csv.DictReader(f)
            for row in reader: yield {norm(k):v for k,v in row.items()}

def county_names(path):
    result={}
    with zipfile.ZipFile(path) as archive:
        with archive.open(archive.namelist()[0]) as raw:
            for row in csv.DictReader(io.TextIOWrapper(raw,encoding="utf-8-sig"),delimiter="|"):
                result[str(row["GEOID"]).zfill(5)]=re.sub(r"\s+(County|Parish|Borough|Census Area|Municipality|city and borough)$","",row["NAME"],flags=re.I)
    return result

def get(row,*names):
    for name in names:
        value=row.get(norm(name))
        if value is not None and str(value).strip()!="": return str(value).strip()
    return ""

p=argparse.ArgumentParser(description="Build GeoList data from the quarterly HUD-USPS ZIP-County bulk file (CSV/XLSX).")
p.add_argument("--input",required=True,type=Path); p.add_argument("--counties",required=True,type=Path); p.add_argument("--output",required=True,type=Path); p.add_argument("--release",default="quarterly bulk file")
a=p.parse_args(); names=county_names(a.counties); records=[]
for row in rows(a.input):
    z=get(row,"ZIP").split(".")[0].zfill(5); fips=get(row,"COUNTY").split(".")[0].zfill(5)
    state=get(row,"USPS_ZIP_PREF_STATE","USPSZIP_PREF_STATE","STATE").upper(); city=get(row,"USPS_ZIP_PREF_CITY","USPSZIP_PREF_CITY","CITY").title()
    if not re.fullmatch(r"\d{5}",z) or not state or not fips: continue
    records.append({"zip":z,"state":state,"stateName":STATE_NAMES.get(state,state),"county":names.get(fips,fips),"countyCode":fips,"primaryCity":city or "Unknown","cities":[city] if city else [],"resRatio":float(get(row,"RES_RATIO") or 0),"busRatio":float(get(row,"BUS_RATIO") or 0),"totRatio":float(get(row,"TOT_RATIO") or 0)})
if len(records)<30000: raise SystemExit(f"Validation failed: only {len(records)} ZIP/county rows were read")
records.sort(key=lambda x:(x["zip"],-x["resRatio"],x["countyCode"]))
unique=len({x["zip"] for x in records}); states=len({x["state"] for x in records})
payload={"meta":{"source":"HUD-USPS ZIP-County Crosswalk","release":a.release,"generatedAt":__import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat(),"recordCount":len(records),"uniqueZipCount":unique,"stateCount":states,"method":"Quarterly bulk file; no runtime API"},"records":records}
a.output.parent.mkdir(parents=True,exist_ok=True); a.output.write_text(json.dumps(payload,separators=(",",":")),encoding="utf-8")
print(json.dumps(payload["meta"],indent=2))
